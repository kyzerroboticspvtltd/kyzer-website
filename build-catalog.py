import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import base64, json

# ── Data ───────────────────────────────────────────────────────────
TABS = [
  {
    "name": "Arduino", "icon": "⚡", "tabColor": "1565C0",
    "category": "electronics", "subcat": "arduino",
    "groups": [
      {
        "label": "Boards", "type": "arduino-boards", "color": "DBEAFE",
        "rows": [
          ["rb-ard-b01","Arduino Uno R3",446,"The most popular Arduino board. ATmega328P with 14 digital I/O and 6 analog inputs."],
          ["rb-ard-b02","Arduino Nano",349,"Compact breadboard-friendly ATmega328P in a tiny footprint."],
          ["rb-ard-b03","Arduino Mega 2560",999,"54 digital I/O and 16 analog inputs on ATmega2560. For large complex projects."],
          ["rb-ard-b04","Arduino Leonardo",649,"ATmega32U4 with built-in USB. Can emulate a mouse or keyboard HID device."],
          ["rb-ard-b05","Arduino Due",1299,"First Arduino with 32-bit ARM Cortex-M3 at 84 MHz and two DAC outputs."],
          ["rb-ard-b06","Arduino Pro Mini 5V",199,"Minimalist ATmega328P at 16 MHz / 5V. Upload via USB-to-TTL adapter."],
          ["rb-ard-b07","Arduino Pro Mini 3.3V",199,"ATmega328P at 8 MHz / 3.3V. Lower power compatible with 3.3V sensors."],
          ["rb-ard-b08","Arduino Micro",799,"ATmega32U4 with native USB in a smaller form factor than the Leonardo."],
          ["rb-ard-b09","Arduino Nano Every",549,"Next-gen Nano with ATMega4809. More flash and RAM on the same pinout."],
          ["rb-ard-b10","Arduino Uno WiFi Rev2",1999,"Uno R3 plus Wi-Fi and Bluetooth via ESP32 co-processor and ATmega4809."],
          ["rb-ard-b11","Arduino MKR WiFi 1010",2499,"32-bit SAMD21 with onboard Wi-Fi and Bluetooth for IoT Cloud projects."],
          ["rb-ard-b12","Arduino MKR Zero",1799,"SAMD21 with SD card and I2S connector for audio and data-logging."],
          ["rb-ard-b13","Arduino Nano 33 IoT",1499,"SAMD21 + NINA-W10 Wi-Fi/BT + LSM6DS3 IMU. Ideal for wearables and IoT."],
          ["rb-ard-b14","Arduino Nano 33 BLE Sense",2199,"nRF52840 + Bluetooth 5 + nine onboard sensors. Machine learning ready."],
          ["rb-ard-b15","Arduino Zero",1599,"SAMD21 Cortex-M0+ with embedded EDBG for full in-circuit debugging."],
        ]
      },
      {
        "label": "Displays", "type": "arduino-displays", "color": "FEF9C3",
        "rows": [
          ["rb-ard-d01","2.4 inch TFT LCD Shield",499,"320x240 colour TFT touchscreen for Uno/Mega. Includes SD card slot."],
          ["rb-ard-d02","3.5 inch TFT LCD Shield",699,"480x320 colour resistive touchscreen for Uno/Mega. Great for GUIs."],
          ["rb-ard-d03","1.8 inch TFT SPI Display",299,"128x160 TFT colour breakout board. Works via hardware or software SPI."],
          ["rb-ard-d04","0.96 inch OLED I2C Display",199,"128x64 monochrome OLED with I2C. Only two signal wires needed."],
          ["rb-ard-d05","1.3 inch OLED I2C Display",249,"128x64 OLED larger than 0.96 inch. SH1106 controller."],
          ["rb-ard-d06","2.42 inch OLED SPI Display",449,"128x64 OLED on SPI for faster refresh. Sharp even in daylight."],
          ["rb-ard-d07","16x2 LCD Display Module",129,"Classic HD44780 character LCD. Add I2C backpack for two-wire control."],
          ["rb-ard-d08","20x4 LCD Display Module",199,"Four rows of 20 characters. Same HD44780 interface as the 16x2."],
          ["rb-ard-d09","7-Segment 4-Digit Display",149,"TM1637 driver with only two wires. Perfect for clocks and counters."],
          ["rb-ard-d10","E-Paper 2.9 inch Display",799,"Tri-colour e-ink via SPI. Retains image with zero power draw."],
        ]
      },
      {
        "label": "Shields", "type": "arduino-shields", "color": "DCFCE7",
        "rows": [
          ["rb-ard-s01","Arduino Sensor Shield V5",349,"Breaks every Uno pin into a 3-pin VCC/GND/SIG header for direct plug-in."],
          ["rb-ard-s02","Arduino Relay Shield 4-Channel",499,"Four 10A/250VAC relays with opto-isolated inputs to protect the MCU."],
          ["rb-ard-s03","Ethernet Shield W5100",599,"Wired Ethernet for Uno/Mega. W5100 chip with RJ45 and onboard SD."],
          ["rb-ard-s04","ESP8266 WiFi Shield",349,"Wi-Fi over serial UART with AT-command firmware pre-loaded."],
          ["rb-ard-s05","Motor Driver Shield L298P",399,"Dual H-bridge for two DC motors or one stepper. 7-12V motor supply."],
          ["rb-ard-s06","CNC Shield V3",299,"Drives up to 4 axes with A4988/DRV8825 drivers for CNC and 3D printers."],
          ["rb-ard-s07","Proto Shield with Mini Breadboard",249,"Stackable shield with a 170-point mini breadboard for quick experiments."],
          ["rb-ard-s08","Data Logger Shield",449,"DS1307 RTC and micro SD on one shield. Timestamps sensor data to CSV."],
          ["rb-ard-s09","Bluetooth Shield HC-05",399,"HC-05 Bluetooth 2.0 with UART passthrough for wireless serial."],
          ["rb-ard-s10","GSM GPRS Shield SIM900",1499,"SIM900 quad-band GSM for SMS voice calls and GPRS data."],
          ["rb-ard-s11","GPS Shield NEO-6M",799,"u-blox NEO-6M GPS with SMA antenna connector and SD card slot."],
          ["rb-ard-s12","Touch Screen Shield 3.2 inch",599,"320x480 resistive touch TFT for Mega with SD and onboard font chip."],
          ["rb-ard-s13","Voice Recognition Shield V3",899,"Up to 80 user-trained offline voice commands. Speaker and mic jack."],
          ["rb-ard-s14","Joystick Shield",199,"Analogue joystick plus 4 buttons mapped to A0/A1 for robot control."],
          ["rb-ard-s15","XBee Shield",549,"Carrier for XBee modules with UART/SPI switch and RSSI LED."],
        ]
      },
      {"label": "Accessories", "type": "arduino-accessories", "color": "FCE7F3", "rows": []},
    ]
  },
  {
    "name": "Raspberry Pi", "icon": "Pi", "tabColor": "C62828",
    "category": "electronics", "subcat": "raspberry-pi",
    "groups": [
      {"label": "Single Board Computers", "type": "rpi-sbc", "color": "FFEBEE", "rows": [
        ["rb-rpi-s01","Raspberry Pi 5 (4GB)",5499,"Latest RPi 5 with 4GB RAM, BCM2712 quad-core 2.4GHz, dual 4K HDMI, USB 3.0, PCIe connector."],
        ["rb-rpi-s02","Raspberry Pi 5 (8GB)",7999,"RPi 5 with 8GB RAM. Best for AI/ML workloads, media servers, and desktop use."],
        ["rb-rpi-s03","Raspberry Pi 4 Model B (4GB)",4499,"Quad-core Cortex-A72 at 1.8GHz with 4GB RAM. Dual 4K HDMI, USB 3.0, Gigabit Ethernet."],
        ["rb-rpi-s04","Raspberry Pi 4 Model B (2GB)",3299,"RPi 4 with 2GB RAM. Great for most projects, servers, and desktop usage."],
        ["rb-rpi-s05","Raspberry Pi Zero 2 W",1299,"Tiny RPi with quad-core Cortex-A53 at 1GHz, 512MB RAM, Wi-Fi, Bluetooth."],
        ["rb-rpi-s06","Raspberry Pi Zero W",799,"Ultra-compact RPi with single-core 1GHz, 512MB RAM, Wi-Fi, and Bluetooth."],
        ["rb-rpi-s07","Raspberry Pi Pico W",349,"RP2040 microcontroller board with Wi-Fi. 264KB RAM, 2MB Flash, 26 GPIO."],
        ["rb-rpi-s08","Raspberry Pi Pico",199,"RP2040 chip with dual Cortex-M0+ at 133MHz. 26 GPIO, 3 ADC channels."],
      ]},
      {"label": "HATs & Add-ons", "type": "rpi-hats", "color": "FCE4EC", "rows": [
        ["rb-rpi-h01","RPi Camera Module 3",2499,"12MP Sony IMX708 autofocus camera. HDR support, 120° FoV wide version available."],
        ["rb-rpi-h02","RPi Camera Module V2",1499,"8MP Sony IMX219 fixed-focus camera for standard RPi CSI connector."],
        ["rb-rpi-h03","Sense HAT",2999,"8x8 RGB LED matrix, joystick, temperature, humidity, pressure, IMU. Used on ISS."],
        ["rb-rpi-h04","Motor Driver HAT",899,"Dual DC motor and stepper driver HAT for RPi. I2C control, up to 12V motors."],
        ["rb-rpi-h05","GPS HAT NEO-6M",1599,"u-blox NEO-6M GPS receiver HAT. PPS output for precise timing applications."],
        ["rb-rpi-h06","PoE+ HAT",1799,"Power over Ethernet HAT for RPi 4/5. IEEE 802.3at PoE+ standard, active cooling."],
      ]},
      {"label": "Accessories", "type": "rpi-accessories", "color": "FFF3E0", "rows": [
        ["rb-rpi-a01","Official RPi 5 Case",799,"Official white/red plastic case for Raspberry Pi 5 with active cooler."],
        ["rb-rpi-a02","Official RPi 4 Case",499,"Official Raspberry Pi 4 plastic case. Available in red/white."],
        ["rb-rpi-a03","Official USB-C Power Supply 5V/3A",699,"Official RPi 4 power adapter, 5V 3A USB-C, white, multi-region plugs."],
        ["rb-rpi-a04","Micro SD Card 32GB Class 10",349,"SanDisk 32GB Class 10 micro SD card. Pre-loaded with Raspberry Pi OS on request."],
        ["rb-rpi-a05","GPIO Ribbon Cable 40-Pin",149,"40-pin ribbon cable to connect RPi GPIO to breadboard or expansion board."],
      ]},
    ]
  },
  {
    "name": "Sensors", "icon": "Snsr", "tabColor": "2E7D32",
    "category": "electronics", "subcat": "sensors",
    "groups": [
      {"label": "Temperature & Humidity", "type": "sensor-temp", "color": "E8F5E9", "rows": [
        ["rb-sns-t01","DHT11 Temperature & Humidity Sensor",99,"Basic digital temp/humidity sensor. ±2°C accuracy, 20–80% RH range. 1-wire protocol."],
        ["rb-sns-t02","DHT22 Temperature & Humidity Sensor",249,"Improved DHT11 with ±0.5°C accuracy, -40 to +80°C range, 0–100% RH."],
        ["rb-sns-t03","DS18B20 Waterproof Temperature Probe",199,"Stainless-steel waterproof 1-wire temperature probe. -55 to +125°C, ±0.5°C."],
        ["rb-sns-t04","BME280 Temp/Humidity/Pressure Module",349,"Bosch BME280 on I2C/SPI breakout. Temp, humidity, and barometric pressure."],
        ["rb-sns-t05","LM35 Temperature Sensor",79,"Analog temperature sensor, 10mV/°C linear output. 0–100°C range."],
        ["rb-sns-t06","MLX90614 IR Temperature Sensor",699,"Non-contact IR thermometer module. -70 to +380°C, I2C interface."],
      ]},
      {"label": "Motion & Distance", "type": "sensor-motion", "color": "F1F8E9", "rows": [
        ["rb-sns-m01","HC-SR04 Ultrasonic Distance Sensor",99,"Ultrasonic ranging module, 2cm–400cm, 3mm accuracy. 5V trigger/echo."],
        ["rb-sns-m02","PIR Motion Sensor HC-SR501",129,"Passive infrared motion detector. 3–7m range, adjustable sensitivity and delay."],
        ["rb-sns-m03","MPU-6050 6-Axis IMU",199,"3-axis gyroscope + 3-axis accelerometer on I2C. ±2/4/8/16g, ±250–2000°/s."],
        ["rb-sns-m04","MPU-9250 9-Axis IMU",449,"9-DOF IMU with gyroscope, accelerometer, and magnetometer on SPI/I2C."],
        ["rb-sns-m05","VL53L0X ToF Distance Sensor",349,"Time-of-flight ranging sensor up to 2m. I2C, works in ambient light."],
        ["rb-sns-m06","ADXL345 3-Axis Accelerometer",199,"3-axis digital accelerometer, ±2/4/8/16g, I2C/SPI. Tap/free-fall detect."],
        ["rb-sns-m07","Reed Switch Module",79,"Magnetic reed switch with digital output. Use with magnets for door/window sensing."],
      ]},
      {"label": "Light & Colour", "type": "sensor-light", "color": "FFFDE7", "rows": [
        ["rb-sns-l01","LDR Photoresistor Module",49,"Light-dependent resistor module with analog output. For basic light sensing."],
        ["rb-sns-l02","BH1750 Ambient Light Sensor",149,"Digital light intensity sensor (lux), I2C. 1–65535 lux range."],
        ["rb-sns-l03","TCS34725 RGB Colour Sensor",349,"RGB and clear light sensing with IR filter. I2C, programmable gain and timing."],
        ["rb-sns-l04","APDS-9960 Gesture Colour Proximity",449,"Combined gesture, proximity, colour, and ambient light sensor on I2C."],
        ["rb-sns-l05","IR Obstacle Avoidance Sensor",89,"Infrared proximity sensor with digital output. Adjustable 2–30cm range."],
      ]},
      {"label": "Gas & Environmental", "type": "sensor-gas", "color": "E0F7FA", "rows": [
        ["rb-sns-g01","MQ-2 Smoke Gas Sensor",149,"Detects LPG, propane, hydrogen, smoke. Analog output, 5V. For gas alarms."],
        ["rb-sns-g02","MQ-135 Air Quality Sensor",149,"Detects NH3, NOx, alcohol, benzene, CO2. Analog output. For air quality monitors."],
        ["rb-sns-g03","MQ-7 Carbon Monoxide Sensor",199,"Highly sensitive CO sensor. 20–2000ppm range. Analog output with heater control."],
        ["rb-sns-g04","CCS811 CO2 & TVOC Sensor",699,"Digital eCO2 and TVOC sensor, I2C. 400–8192ppm CO2, indoor air quality."],
        ["rb-sns-g05","Rain/Water Drop Sensor",99,"Conductive rain detection sensor with analog and digital output. 5V."],
        ["rb-sns-g06","Soil Moisture Sensor",99,"Resistive soil moisture probe with analog output. For plant watering automation."],
      ]},
    ]
  },
  {
    "name": "Motors", "icon": "Mtr", "tabColor": "E65100",
    "category": "electronics", "subcat": "motors",
    "groups": [
      {"label": "DC Motors", "type": "motor-dc", "color": "FBE9E7", "rows": [
        ["rb-mtr-d01","DC Geared Motor 6V 60RPM",249,"Metal gear DC motor, 6V, 60RPM, high torque. For robot arms and heavy-load robots."],
        ["rb-mtr-d02","DC Geared Motor 6V 120RPM",249,"6V 120RPM gear motor. Good balance of speed and torque for wheeled robots."],
        ["rb-mtr-d03","DC Geared Motor 12V 200RPM",399,"12V 200RPM high-speed gear motor for robotics carts and conveyor projects."],
        ["rb-mtr-d04","BO Motor (Battery Operated) 3V 150RPM",129,"Lightweight BO motor ideal for mini line-follower and sumo robots."],
        ["rb-mtr-d05","N20 Micro Metal Gear Motor 6V 100RPM",349,"Tiny N20 micro metal gear motor, 6V, 100RPM. For compact robot platforms."],
      ]},
      {"label": "Servo Motors", "type": "motor-servo", "color": "FFF3E0", "rows": [
        ["rb-mtr-s01","SG90 Micro Servo 9g",149,"9g micro servo, 180° rotation, 1.8kg·cm torque. Ideal for small arms and fins."],
        ["rb-mtr-s02","MG996R Metal Gear Servo",349,"Metal gear servo, 180°, 9.4–11kg·cm torque. For robotic arms and steering."],
        ["rb-mtr-s03","MG90S Metal Gear Micro Servo",199,"Metal gear mini servo, 180°, 2.2kg·cm. Stronger than SG90 in same size."],
        ["rb-mtr-s04","DS3218 270° Digital Servo",799,"High-torque digital servo, 270° rotation, 20kg·cm. For industrial-grade arms."],
        ["rb-mtr-s05","360° Continuous Rotation Servo",249,"Continuous rotation servo with speed and direction control via PWM."],
      ]},
      {"label": "Stepper Motors", "type": "motor-stepper", "color": "FFF8E1", "rows": [
        ["rb-mtr-st01","28BYJ-48 Stepper Motor + ULN2003 Driver",199,"5V unipolar stepper with driver board. 512 steps/rev, 1:64 reduction for precision."],
        ["rb-mtr-st02","NEMA 17 Stepper Motor 40Ncm",699,"NEMA 17 bipolar stepper, 1.8° per step, 40Ncm torque. Standard for 3D printers."],
        ["rb-mtr-st03","NEMA 17 Stepper Motor 59Ncm",899,"High-torque NEMA 17, 59Ncm. For CNC machines and heavier 3D printer axes."],
        ["rb-mtr-st04","NEMA 23 Stepper Motor 130Ncm",1899,"NEMA 23 bipolar stepper, 1.8°/step, 130Ncm. For CNC mills and lathes."],
      ]},
      {"label": "Motor Drivers", "type": "motor-driver", "color": "F9FBE7", "rows": [
        ["rb-mtr-dr01","L298N Dual H-Bridge Motor Driver",149,"Classic L298N driver module. 2 DC motors or 1 stepper, up to 2A per channel, 7–46V."],
        ["rb-mtr-dr02","L293D Motor Driver IC",79,"Quad half-H-bridge IC, 600mA per channel, 4.5–36V. DIP-16 package for breadboard."],
        ["rb-mtr-dr03","A4988 Stepper Motor Driver",149,"A4988 microstepping driver up to 1/16 step. 35V, 2A. Standard for 3D printers."],
        ["rb-mtr-dr04","DRV8825 Stepper Motor Driver",199,"Up to 1/32 microstepping, 45V, 2.5A. Quieter than A4988, ideal for CNC."],
        ["rb-mtr-dr05","TB6612FNG Dual Motor Driver",199,"Dual H-bridge driver, 1.2A continuous per channel. Smaller and more efficient than L298N."],
        ["rb-mtr-dr06","BTS7960 43A High-Power Motor Driver",499,"High-current half-bridge driver module, 43A peak. For heavy-duty DC motors."],
      ]},
    ]
  },
  {
    "name": "Components", "icon": "Comp", "tabColor": "4527A0",
    "category": "electronics", "subcat": "components",
    "groups": [
      {"label": "Resistors & Capacitors", "type": "comp-passive", "color": "EDE7F6", "rows": [
        ["rb-cmp-p01","Resistor Kit 600pcs (30 values)",199,"600-piece resistor assortment, 1Ω–1MΩ, 1/4W 5% tolerance. Sorted in compartment box."],
        ["rb-cmp-p02","Capacitor Kit 300pcs (30 values)",249,"Electrolytic and ceramic capacitor assortment from 1pF to 1000µF."],
        ["rb-cmp-p03","LED Assortment Kit 100pcs (5 colours)",99,"5mm LEDs: red, green, blue, yellow, white, 20 pcs each. Basic indicator LEDs."],
        ["rb-cmp-p04","Push Button Tactile Switch Pack (20pcs)",79,"6x6mm through-hole tactile switches. 4-pin. For Arduino shields and PCBs."],
        ["rb-cmp-p05","Potentiometer 10K Linear (5pcs)",99,"10kΩ single-turn rotary potentiometer, B10K. 3-pin, panel or PCB mount."],
        ["rb-cmp-p06","Zener Diode Kit (5V, 12V, 24V assorted)",149,"Assorted zener diodes for voltage regulation and protection circuits."],
      ]},
      {"label": "ICs & Modules", "type": "comp-ic", "color": "E8EAF6", "rows": [
        ["rb-cmp-i01","NE555 Timer IC",29,"Classic 555 timer in DIP-8. For oscillators, pulse generation, and timing circuits."],
        ["rb-cmp-i02","LM741 Op-Amp IC",39,"General-purpose single op-amp in DIP-8. Amplification, comparators, filters."],
        ["rb-cmp-i03","74HC595 Shift Register IC",49,"8-bit serial-to-parallel shift register. Expand Arduino outputs using 3 pins."],
        ["rb-cmp-i04","PCF8574 I2C GPIO Expander",99,"8-bit I/O expander for I2C bus. Add 8 GPIO pins via just two wires."],
        ["rb-cmp-i05","AMS1117 3.3V Voltage Regulator",39,"SOT-223 LDO regulator, 800mA, 3.3V output. Popular on ESP and RPi projects."],
        ["rb-cmp-i06","ESP-01 ESP8266 Wi-Fi Module",249,"Tiny ESP8266 Wi-Fi module with 2 GPIO. Add Wi-Fi to any Arduino via serial."],
        ["rb-cmp-i07","HC-05 Bluetooth Module",349,"Classic HC-05 Bluetooth 2.0 serial module. Master/slave mode via AT commands."],
        ["rb-cmp-i08","nRF24L01+ 2.4GHz RF Module",199,"2.4GHz SPI radio module, 100m+ range, 250kbps–2Mbps. Pair for wireless comm."],
      ]},
      {"label": "Connectors & Wires", "type": "comp-connector", "color": "E3F2FD", "rows": [
        ["rb-cmp-c01","Jumper Wire Kit M-M/M-F/F-F (120pcs)",149,"40pcs each male-male, male-female, female-female jumper wires, 20cm."],
        ["rb-cmp-c02","Breadboard 830 Tie Points",199,"Full-size solderless breadboard, 830 tie-points, with binding posts."],
        ["rb-cmp-c03","Mini Breadboard 400 Tie Points",99,"Half-size 400 tie-point breadboard. Ideal for compact circuit prototyping."],
        ["rb-cmp-c04","USB-A to USB-B Cable 1m",99,"Standard USB cable for Arduino Uno/Mega programming and power."],
        ["rb-cmp-c05","USB-A to Micro USB Cable 1m",79,"Micro USB cable for Arduino Nano, Pro Mini adapter, and RPi power."],
        ["rb-cmp-c06","Dupont Connector Kit with Crimping Tool",349,"Male/female Dupont housings, crimp pins, and crimping tool. Make custom cables."],
        ["rb-cmp-c07","XT60 Connector Pair (Male + Female)",99,"High-current XT60 connector pair rated 60A. Standard for LiPo battery packs."],
        ["rb-cmp-c08","18650 Battery Holder 2-Cell",79,"2x18650 lithium cell battery holder with leads. 7.4V nominal output."],
      ]},
    ]
  },
  {
    "name": "Tools", "icon": "Tool", "tabColor": "37474F",
    "category": "electronics", "subcat": "tools",
    "groups": [
      {"label": "Soldering", "type": "tool-solder", "color": "ECEFF1", "rows": [
        ["rb-tl-sl01","Soldering Iron 25W",399,"Basic 25W pencil soldering iron. 400°C tip temperature. For PCB and wire work."],
        ["rb-tl-sl02","Soldering Iron Station 60W",1499,"Adjustable 60W soldering station with digital temperature control 200–480°C."],
        ["rb-tl-sl03","Solder Wire 60/40 250g",299,"60/40 tin-lead solder, 0.8mm diameter, rosin flux core. 250g spool."],
        ["rb-tl-sl04","Desoldering Pump / Solder Sucker",199,"Spring-loaded desoldering pump for removing excess solder from PCB joints."],
        ["rb-tl-sl05","Flux Paste 50g",149,"Water-washable flux paste in a jar. Improves solder flow and reduces oxidation."],
        ["rb-tl-sl06","PCB Helping Hands with Magnifier",399,"Third-hand tool with alligator clips and 5x magnifying glass for PCB work."],
      ]},
      {"label": "Measuring", "type": "tool-measure", "color": "F5F5F5", "rows": [
        ["rb-tl-ms01","Digital Multimeter",799,"Auto-ranging digital multimeter, DC/AC voltage, current, resistance, continuity."],
        ["rb-tl-ms02","USB Logic Analyser 8-Channel",1499,"8-channel 24MHz USB logic analyser. Works with Sigrok/PulseView. I2C/SPI/UART decode."],
        ["rb-tl-ms03","Oscilloscope DSO138 DIY Kit",1999,"DIY 200kHz digital oscilloscope kit with 2.4-inch LCD. Single-channel, 1Msps."],
        ["rb-tl-ms04","LCR Meter Component Tester",999,"Auto-detect transistors, resistors, capacitors, inductors, diodes on a graphic LCD."],
        ["rb-tl-ms05","Bench Power Supply 30V 5A",3499,"Variable bench PSU, 0–30V 0–5A, dual display, over-current protection."],
      ]},
      {"label": "Hand Tools", "type": "tool-hand", "color": "FAFAFA", "rows": [
        ["rb-tl-ht01","Precision Screwdriver Set 25pcs",499,"25-piece magnetic screwdriver set with Phillips, flat, Torx, hex bits."],
        ["rb-tl-ht02","Wire Stripper & Cutter",299,"Adjustable wire stripper for 10–24 AWG wire with built-in cutter."],
        ["rb-tl-ht03","Needle-Nose Pliers",199,"Long-nose ESD-safe pliers for bending component leads and handling SMD parts."],
        ["rb-tl-ht04","Diagonal Flush Cutters",149,"Micro flush cutter for clean component lead trimming after soldering."],
        ["rb-tl-ht05","Anti-Static ESD Wrist Strap",99,"ESD wrist strap with 1MΩ resistor and alligator clip for static-safe PCB work."],
        ["rb-tl-ht06","Hot Glue Gun 20W with Sticks",299,"Low-temperature mini glue gun with 10 glue sticks. For enclosure and wire tack."],
      ]},
    ]
  },
  {
    "name": "Drones", "icon": "Drn", "tabColor": "1B5E20",
    "category": "drone", "subcat": "drone",
    "groups": [
      {"label": "Frames", "type": "drone-frame", "color": "E8F5E9", "rows": [
        ["rb-drn-f01","F450 Quadcopter Frame 450mm",799,"Classic F450 glass-fibre quad frame with PCB bottom plate for power distribution."],
        ["rb-drn-f02","Q250 Racing Frame 250mm",999,"Compact 250mm carbon-fibre FPV racing frame. Low profile, strong arm design."],
        ["rb-drn-f03","S500 Hexacopter Frame 500mm",1499,"S500 glass-fibre hexacopter frame, retractable landing gear, PCB power plate."],
        ["rb-drn-f04","X500 V2 Quadcopter Frame 500mm",2499,"All-carbon X500 V2 frame compatible with Holybro Pixhawk stacks."],
      ]},
      {"label": "Complete Drones", "type": "drone-complete", "color": "F1F8E9", "rows": [
        ["rb-drn-c01","Survey Mapping Drone S1000",85000,"Fixed-wing VTOL mapping drone. 60-min flight time, RTK GPS, 1km range."],
        ["rb-drn-c02","FPV Racing Drone 5-Inch",18000,"Ready-to-fly 5-inch FPV racer with F7 FC, 45A ESC, and digital FPV system."],
        ["rb-drn-c03","Agricultural Spray Drone 10L",250000,"10L tank hexacopter for precision crop spraying. 15-min per 10 acres coverage."],
      ]},
      {"label": "Parts & Components", "type": "drone-parts", "color": "FFFDE7", "rows": [
        ["rb-drn-p01","A2212 1000KV Brushless Motor",599,"2212 1000KV brushless outrunner for 450-class quads. 250W, fits 10-inch props."],
        ["rb-drn-p02","A2212 2200KV Brushless Motor",599,"2212 2200KV brushless motor for 250mm racers and 330-class quads."],
        ["rb-drn-p03","30A Brushless ESC",399,"30A UBEC ESC with BEC output, compatible with most 2-4S LiPo packs."],
        ["rb-drn-p04","FlySky FS-i6X 10CH Transmitter + Receiver",3499,"10-channel 2.4GHz digital transmitter with FS-iA10B receiver for RC aircraft."],
        ["rb-drn-p05","Flight Controller F405 Wing",2999,"F405-based FC with OSD, barometer, and BEC. For fixed-wing and quad builds."],
        ["rb-drn-p06","LiPo Battery 3S 2200mAh 25C",1299,"3S 11.1V 2200mAh 25C LiPo. Standard battery for 450-class quadcopters."],
        ["rb-drn-p07","LiPo Battery 4S 1500mAh 50C",1499,"4S 14.8V 1500mAh 50C LiPo. For 5-inch FPV racing drones."],
        ["rb-drn-p08","1045 Propeller Set (2CW + 2CCW)",199,"10-inch 4.5-pitch carbon propellers for 450-class frames. One set of 4."],
        ["rb-drn-p09","5045 Tri-Blade FPV Propeller Set",249,"5-inch 3-blade FPV racing propellers. One set of 4 (2CW + 2CCW)."],
        ["rb-drn-p10","GPS Module M8N with Compass",1499,"u-blox M8N GPS + HMC5883L compass on a pole mount for ArduPilot/PX4."],
      ]},
    ]
  },
  {
    "name": "3D Printing", "icon": "3DP", "tabColor": "880E4F",
    "category": "3dprint", "subcat": "3dprint",
    "groups": [
      {"label": "Printers", "type": "printer", "color": "FCE4EC", "rows": [
        ["rb-3dp-pr01","Creality Ender 3 V3 SE",14999,"Entry-level FDM printer. Auto bed levelling, 250mm/s print speed, 220x220x250mm."],
        ["rb-3dp-pr02","Bambu Lab A1 Mini",34999,"Multi-colour capable FDM printer with auto-calibration. 180x180x180mm build volume."],
        ["rb-3dp-pr03","Elegoo Saturn 3 Ultra MSLA",54999,"12K mono LCD resin printer. 218x123x220mm, 0.019mm XY resolution."],
      ]},
      {"label": "Filament", "type": "filament", "color": "FFF0F5", "rows": [
        ["rb-3dp-fl01","PLA+ Filament 1kg — Black",1299,"High-quality PLA+ filament, 1.75mm, 1kg spool. Easy to print, low warping."],
        ["rb-3dp-fl02","PLA+ Filament 1kg — White",1299,"White PLA+ filament, 1.75mm, 1kg. Good for painting and finishing."],
        ["rb-3dp-fl03","PETG Filament 1kg — Transparent",1499,"Clear PETG filament, 1.75mm, 1kg. Food-safe, chemical resistant, low shrink."],
        ["rb-3dp-fl04","TPU Filament 95A 1kg — Black",1799,"Flexible 95A TPU filament, 1.75mm. For phone cases, gaskets, grippers."],
        ["rb-3dp-fl05","ABS Filament 1kg — Grey",1199,"ABS filament 1.75mm 1kg. High-temperature resistance. Requires enclosure."],
        ["rb-3dp-fl06","ASA Filament 1kg — Black",1599,"UV and weather-resistant ASA, 1.75mm, 1kg. For outdoor drone and RC parts."],
      ]},
      {"label": "Accessories", "type": "print-accessories", "color": "F8BBD0", "rows": [
        ["rb-3dp-ac01","Brass Nozzle 0.4mm MK8",99,"Standard MK8 brass nozzle, 0.4mm for Ender/Creality printers. Pack of 5."],
        ["rb-3dp-ac02","Hardened Steel Nozzle 0.4mm",299,"Hardened steel MK8 nozzle for abrasive filaments (carbon fibre, glow)."],
        ["rb-3dp-ac03","Bed Adhesion PEI Spring Sheet 235x235",799,"Magnetic PEI spring steel sheet for Ender 3 series. Easy print removal."],
        ["rb-3dp-ac04","Spatula Print Removal Tool",149,"Stiff metal spatula for safe print removal from glass or PEI beds."],
        ["rb-3dp-ac05","Filament Dry Box with Roller",699,"Sealed filament storage box with humidity card and PTFE outlet tube."],
      ]},
    ]
  },
  {
    "name": "Prototyping", "icon": "Proto", "tabColor": "F57F17",
    "category": "proto", "subcat": "prototyping",
    "groups": [
      {"label": "Breadboards & PCBs", "type": "proto-breadboard", "color": "FFFDE7", "rows": [
        ["rb-prt-b01","Breadboard 830 Tie Points",199,"Full-size 830-point solderless breadboard with binding posts. Standard for prototyping."],
        ["rb-prt-b02","Double-Sided PCB 5x7cm (5pcs)",99,"5-pack of 5x7cm double-sided perf boards. 2.54mm hole pitch for DIP components."],
        ["rb-prt-b03","Double-Sided PCB 10x15cm (2pcs)",149,"2-pack of 10x15cm double-sided universal PCBs. Ideal for larger project boards."],
        ["rb-prt-b04","SMD Prototype PCB 5x7cm (5pcs)",129,"5-pack of SMD-capable PCBs with 0805 pads and SOT-23 footprints."],
      ]},
      {"label": "Development Kits", "type": "proto-kit", "color": "FFF8E1", "rows": [
        ["rb-prt-k01","Arduino Starter Kit with Components",1499,"Arduino Uno R3, breadboard, jumper wires, LEDs, resistors, sensors — complete starter box."],
        ["rb-prt-k02","ESP32 IoT Development Kit",999,"ESP32 dev board, DHT22, relay, OLED, jumper wires — for IoT and Wi-Fi projects."],
        ["rb-prt-k03","Raspberry Pi Pico Starter Kit",799,"RPi Pico W, breadboard, LEDs, buttons, resistors, micro USB cable, and guide."],
        ["rb-prt-k04","Robot Car Chassis Kit (4WD)",1299,"4-wheel drive chassis with DC motors, wheels, encoders, and motor driver module."],
        ["rb-prt-k05","Line Follower Robot Kit",1499,"Complete line-follower kit with IR sensors, L298N, Arduino Uno, and chassis."],
      ]},
      {"label": "Enclosures", "type": "proto-enclosure", "color": "FFF9C4", "rows": [
        ["rb-prt-e01","ABS Project Box 100x60x25mm",149,"Black ABS enclosure for small electronics projects. Screw-top lid, 100x60x25mm."],
        ["rb-prt-e02","ABS Project Box 150x80x35mm",199,"Medium ABS project enclosure with flanged lid, 150x80x35mm."],
        ["rb-prt-e03","DIN Rail Enclosure 100mm",299,"Snap-on ABS DIN rail enclosure for panel mounting. Fits standard rail."],
        ["rb-prt-e04","Raspberry Pi 4 Acrylic Case",299,"Clear acrylic layered case for RPi 4 with GPIO access, 4 standoffs, fan mount."],
        ["rb-prt-e05","Arduino Uno Acrylic Case",149,"Clear acrylic case for Arduino Uno with USB and power cutouts."],
      ]},
    ]
  },
]

HEADERS = ["ID", "Name", "Price (Rs)", "Description", "Button", "Visible", "category", "subcat", "type"]
COL_WIDTHS = [22, 32, 12, 58, 12, 10, 16, 16, 22]

def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

for tab in TABS:
    ws = wb.create_sheet(title=tab["name"])
    ws.sheet_properties.tabColor = tab["tabColor"]

    # ── Header row ──────────────────────────────────────────────
    ws.append(HEADERS)
    hdr_fill = hex_fill(tab["tabColor"])
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col_i, width in enumerate(COL_WIDTHS, 1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────
    row_num = 2
    dv_button  = DataValidation(type="list", formula1='"buy,enquire"', showDropDown=False)
    dv_visible = DataValidation(type="list", formula1='"yes,no"',      showDropDown=False)
    ws.add_data_validation(dv_button)
    ws.add_data_validation(dv_visible)

    for g in tab["groups"]:
        # Section header
        ws.append(["  " + g["label"], "", "", "", "", "", "", "", ""])
        sec_fill = hex_fill(g["color"])
        sec_font = Font(bold=True, size=10, color="222222")
        sec_border = thin_border("AAAAAA")
        for col_i in range(1, 10):
            cell = ws.cell(row=row_num, column=col_i)
            cell.fill = sec_fill
            cell.font = sec_font
            cell.border = sec_border
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        data_rows = g["rows"] if g["rows"] else [["", "", "", ""]]
        for r in data_rows:
            row_data = [r[0], r[1], r[2] if len(r) > 2 else "", r[3] if len(r) > 3 else "",
                        "buy", "yes", tab["category"], tab["subcat"], g["type"]]
            ws.append(row_data)

            prod_fill = hex_fill(g["color"])
            meta_fill = hex_fill("F8F8F8")
            for col_i in range(1, 10):
                cell = ws.cell(row=row_num, column=col_i)
                cell.fill = prod_fill if col_i <= 6 else meta_fill
                cell.alignment = Alignment(vertical="center", wrap_text=(col_i == 4))
                if col_i >= 7:
                    cell.font = Font(color="AAAAAA", italic=True, size=9)
                if r[0]:  # only add validation on real rows
                    if col_i == 5:
                        dv_button.add(cell)
                    elif col_i == 6:
                        dv_visible.add(cell)

            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # Hide meta columns (G=7, H=8, I=9)
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    # Auto-filter on header
    ws.auto_filter.ref = f"A1:{get_column_letter(9)}{row_num - 1}"

# ── Save & base64 encode ──────────────────────────────────────────────
OUT = r"c:\Desktop\Kyzer website\kyzer-catalog.xlsx"
wb.save(OUT)
with open(OUT, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()

with open(r"c:\Desktop\Kyzer website\catalog-b64.txt", "w") as f:
    f.write(b64)

print("DONE", len(b64), "chars")
